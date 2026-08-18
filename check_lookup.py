# -*- coding: utf-8 -*-
"""
Сверка парсера с эталонами из рабочих книг.

Запуск:
    ./.venv/bin/python check_lookup.py            # все эталоны
    ./.venv/bin/python check_lookup.py --offline  # без сети, только разбор размеров
    ./.venv/bin/python check_lookup.py <url>      # разовая проверка любой ссылки

Эталоны взяты из реальных спецификаций: колонки J/K/L (габариты),
R (объём) и текст описания. Скрипт показывает, что сошлось, а что нет —
и не притворяется, что расхождение это норма.
"""

from __future__ import annotations

import json
import os
import sys

from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

import product_lookup as pl  # noqa: E402

OK, BAD, WARN = "✓", "✗", "~"

# --- Эталоны разбора размеров: строка из книги -> Д, Г, В --------------
DIMS_CASES = [
    ("130x47x45Н",             "Банкетка",              (130, 47, 45),  "LONGHI ARIANA, R16"),
    ("90x84x80Н",              "Кресло",                (90, 84, 80),   "FLOU MADAME BUTTERFLY, R18"),
    ("D25/31x125Н",            "Торшер",                (31, 31, 125),  "VENICEM CIRCLE, R20"),
    ("202x241x36/92Н",         "Кровать",               (202, 241, 92), "TRUSSARDI VIBES, R22"),
    ("66x60x44Н SPECIAL",      "Тумбочка прикроватная", (66, 60, 44),   "TRUSSARDI COMFY, R23"),
    # Нотации живых сайтов: в книгах их нет, а ошибаются они дороже всего.
    # 43 1/4" = 110 см, 47 1/4" = 120 см, 25 5/8" = 65 см — это дюймовые
    # двойники диаметров; единственный несдиаметральный сантиметр — высота 75.
    ('Ø110 - 120, 43 1/4" - 47 1/4", 75 cm, Ø65, 25 5/8"',
                               "Стол",                  (120, 120, 75), "PORADA INFINITY, сайт"),
    ("135 cm, 100 cm, 64 cm",  "Кресло",                (135, 100, 64), "HENGE RADICAL, сайт"),
    ("Ø 120 cm, H 75 cm",      "Стол",                  (120, 120, 75), "круглый стол с меткой H"),
    ('200x52x80 cm (78 3/4"x20 1/2"x31 1/2")',
                               "Комод",                 (200, 52, 80),  "дюймы в скобках"),
    # Подписи словами — так пишет BAROVIER, и без них два числа неразличимы
    ("Height 60 cm, Diameter 93 cm",
                               "Люстра",                (93, 93, 60),   "BAROVIER METROPOLIS"),
    ("Altezza 120 cm, Diametro 127 cm",
                               "Люстра",                (127, 127, 120), "то же по-итальянски"),
    ("72x76x75H cm",           "Кресло",                (72, 76, 75),   "BENTLEY MERE, сайт"),
    ("300x90x75(h)cm",         "Стол",                  (300, 90, 75),  "HENGE SISMA, пометка в скобках"),
    # Тот же BAROVIER, но извлечение подписало размеры после числа
    ("90 cm (высота), 127 cm (диаметр)",
                               "Люстра",                (127, 127, 90), "подпись в скобках"),
    # PORADA пишет высоту строчной «h» после списка диаметров
    ("Ø130 - 140 - 150 - 160 h 75",
                               "Стол",                  (160, 160, 75), "PORADA, техлист"),
    ("120 x 180 - 200 h 75",   "Стол",                  (120, 180, 75), "PORADA, овал"),
]

# --- Эталоны объёма: габариты -> м³ по колонке R книги -----------------
VOLUME_CASES = [
    ((130, 47, 45),  0.5, "R16"),
    ((90, 84, 80),   1.0, "R18"),
    ((31, 31, 125),  0.2, "R20"),
    ((202, 241, 92), 6.8, "R22"),
    # Не из книги: до правки высотой считался диаметр 120 и выходило 2,2 м³.
    ((120, 120, 75), 1.7, "PORADA INFINITY"),
]

# --- Тип по разделу сайта: адрес -> тип ('' = раздел неоднозначен) ------
URL_TYPE_CASES = [
    ("https://www.emmemobili.it/en/prodotti/contenitori/fractal", "Комод"),
    ("https://www.misuraemme.it/en/products/baltimora-bed",       "Кровать"),
    ("https://www.brand.it/en/products/armchairs/soft",           "Кресло"),
    ("https://www.brand.it/en/products/chairs/pin",               "Стул"),
    ("https://www.brand.it/prodotti/comodini/x",     "Тумбочка прикроватная"),
    ("https://www.porada.it/en/products/side-coffee-tables",      "Стол"),
    # Живой промах: настольная лампа BAROVIER стала «Столом» — голый
    # токен «table» совпадал с table-lamps и перекрывал извлечение.
    ("https://www.barovier.com/en/table-lamps/aurora",  "Настольная лампа"),
    ("https://www.barovier.com/en/chandeliers/metropolis",        "Люстра"),
    ("https://www.barovier.com/en/floor-lamps/lume",              "Торшер"),
    ("https://brand.it/lampade-da-tavolo/x",            "Настольная лампа"),
    # «chairs» лежит внутри «armchairs» — раздел смешанный, тип не навязываем
    ("https://www.henge07.com/products/sofas-and-armchairs/cohiba/", ""),
    ("https://www.porada.it/prodotto/infinity",                   ""),
    # У света тип задаёт способ монтажа, а не раздел
    ("https://flos.com/en/us/ic-lights-floor/M-ic-lights-floor.html", ""),
]

# --- Тип из извлечения -> тип из нашего списка --------------------------
# Модель отвечает не только словом из списка: LONGHI вернул «Банкетка / пуф»,
# HENGE — «Стол обеденный». Оба уезжали в «Другое», хотя нужное слово в них
# есть. «Диван-кровать» остаётся «Другим» намеренно: это не диван.
TYPE_NORM_CASES = [
    ("Кресло",         "Кресло"),
    ("Банкетка / пуф", "Банкетка"),
    ("Стол обеденный", "Стол"),
    ("Пуф",            "Пуф"),
    ("Диван-кровать",  "Другое"),
    ("",               ""),
]

# --- Живые ссылки: что ожидаем увидеть ---------------------------------
LIVE_CASES = [
    {
        "url": "https://www.venicem.com/product/circle-floor/",
        "note": "есть техлист — должна собраться полная карточка",
        "expect": {"type_ru": "Торшер", "width_cm": 31, "depth_cm": 31, "height_cm": 125},
    },
    {
        "url": "https://www.longhi.it/en-us/products/bench-pouf/arianna",
        "note": "техлиста нет — габариты должны остаться пустыми, а не выдуманными",
        "expect": {"type_ru": "Банкетка"},
    },
]


def check_dims() -> tuple[int, int]:
    print("\n РАЗБОР РАЗМЕРОВ (без сети)")
    print(" " + "-" * 74)
    good = 0
    for raw, type_ru, expected, source in DIMS_CASES:
        w, d, h, sure = pl.parse_dims(raw, type_ru)
        got = tuple(int(x) if x else 0 for x in (w, d, h))
        hit = got == expected
        good += hit
        mark = OK if hit else BAD
        note = "" if sure else "  (оси угаданы, не по пометке H)"
        print(f"  {mark} {raw:22} -> {got}{note}")
        if not hit:
            print(f"      ожидали {expected} — {source}")
    return good, len(DIMS_CASES)


def check_volume() -> tuple[int, int]:
    print("\n ОБЪЁМ ПО ФОРМУЛЕ КНИГИ")
    print(" " + "-" * 74)
    good = 0
    for (w, d, h), expected, source in VOLUME_CASES:
        got = pl.volume_m3(w, d, h)
        hit = abs(got - expected) < 0.05
        good += hit
        print(f"  {OK if hit else BAD} {w}x{d}x{h:<6} -> {got:.1f} м³   в книге {expected} ({source})")
    return good, len(VOLUME_CASES)


def check_dims_grounding() -> tuple[int, int]:
    """Габариты, которых нет в источнике, обязаны отбиваться.

    Случай из жизни: страница FENDI CASA не публикует размеров вовсе,
    а извлечение вернуло правдоподобные «83 / 81 / 88» — и карточка
    посчитала по ним объём.
    """
    print("\n СВЕРКА ГАБАРИТОВ С ИСТОЧНИКОМ")
    print(" " + "-" * 74)
    cases = [
        ("высота: 83 см, ширина: 81 см, глубина: 88 см",
         "Кресло Cleo, кожа и дерево. Фото и описание без размеров.",
         False, "выдумка на странице без размеров"),
        ("72x76x75H cm", "Mere armchair 72x76x75H cm, fabric or leather",
         True, "есть в источнике"),
        ("Height 60 cm; Diameter 93 cm", "Metropolis Height 60 cm Diameter 93 cm Weight 15 kg",
         True, "подписи словами"),
        ("202x241x92H.", "…165x200 - 202x241x92H. - H.B.36cm…",
         True, "из техлиста"),
        ("", "любой текст", True, "пустая строка не проверяется"),
    ]
    good = 0
    for dims, source, expected, note in cases:
        got = pl._dims_grounded(dims, source)
        hit = got == expected
        good += hit
        print(f"  {OK if hit else BAD} {(dims or '(пусто)')[:44]:46} -> "
              f"{'принято' if got else 'отбито':8} {note}")
    return good, len(cases)


def check_header_roundtrip() -> tuple[int, int]:
    """Шапка проекта переживает выгрузку и обратный разбор.

    Две половины писались порознь: выгрузка складывает книгу, печать её
    разбирает. Пока шапки не было, документ выходил с пустым номером и
    прочерками — и заметить это можно было только глазами на печати.
    """
    import book_export
    import spec_parser
    print("\n ШАПКА: ВЫГРУЗКА -> ПЕЧАТЬ")
    print(" " + "-" * 74)
    head = {"number": "2867/3", "contract": "2867",
            "buyer": "Иванов И. И.", "date": "12.03.2026"}
    position = [{"brand": "VENICEM", "model": "CIRCLE", "qty": 1,
                 "description": "CIRCLE\nТоршер"}]
    spec = spec_parser.parse(book_export.build(position, header=head))
    good = 0
    for key, got in (("number", spec.number), ("contract", spec.contract),
                     ("buyer", spec.buyer), ("date", spec.date)):
        hit = got == head[key]
        good += hit
        print(f"  {OK if hit else BAD} {key:9} -> {got!r}")

    # Пустая шапка не должна порождать «Спецификация № к Договору».
    bare = spec_parser.parse(book_export.build(position))
    hit = not (bare.number or bare.buyer or bare.date)
    good += hit
    print(f"  {OK if hit else BAD} без шапки документ остаётся пустым, "
          f"а не «Спецификация №»")
    return good, 5


def check_download_headers() -> tuple[int, int]:
    """Заголовки выгрузки должны кодироваться latin-1.

    Прод падал уже после сборки файла: кириллица в filename= роняла
    ответ, браузер получал обрыв и показывал общее «не удалось собрать
    файл». Локальный сервер это прощал, поэтому проверка тут.
    """
    import importlib
    import os
    os.environ.setdefault("AURRUM_PASSWORD", "проверка")
    os.environ.setdefault("AURRUM_SECRET_KEY", "x" * 32)
    import app as flask_app
    importlib.reload(flask_app)

    print("\n ЗАГОЛОВКИ ВЫГРУЗКИ")
    print(" " + "-" * 74)
    client = flask_app.app.test_client()
    with client.session_transaction() as sess:
        sess["authorized"] = True
    got = client.post("/project/export", json={"positions": [
        {"brand": "VENICEM", "model": "CIRCLE", "qty": 1, "list_price": 2550}]})

    checks = []
    checks.append(("ответ отдан", got.status_code == 200))
    for name, value in got.headers.items():
        try:
            value.encode("latin-1")
            ok = True
        except UnicodeEncodeError:
            ok = False
        if not ok:
            checks.append((f"{name} кодируется latin-1", False))
    checks.append(("все заголовки кодируются latin-1",
                   all(hit for _, hit in checks)))
    disposition = got.headers.get("Content-Disposition", "")
    checks.append(("настоящее имя приходит в filename*",
                   "filename*=UTF-8''" in disposition))

    good = 0
    for label, hit in checks:
        good += hit
        print(f"  {OK if hit else BAD} {label}")
    return good, len(checks)


def check_library() -> tuple[int, int]:
    """Библиотека: опознаватель, поиск и вход под замком.

    Сеть здесь не нужна: проверяем то, что решает судьбу карточки, —
    как из бренда и модели получается имя файла (ошибка здесь затирает
    чужую карточку) и находится ли товар по своим словам.
    """
    import library
    print("\n БИБЛИОТЕКА")
    print(" " + "-" * 74)
    items = [
        {"id": "porada-infinity", "brand": "Porada", "model": "Infinity",
         "type_ru": "Стол", "summary_ru": "Основание из массива дерева.",
         "finishes": [{"role_ru": "Дерево", "material": "Ashwood"}]},
        {"id": "barovier-toso-aurora", "brand": "Barovier&Toso", "model": "Aurora",
         "type_ru": "Настольная лампа",
         "finishes": [{"role_ru": "Стекло", "material": "Light Pink/Crystal",
                       "code": "AE"}]},
    ]
    find = lambda q: [it["id"] for it in library.search(items, q)]
    checks = [
        ("бренд и модель -> имя файла",
         library.slug("Barovier&Toso", "Aurora") == "barovier-toso-aurora"),
        # Без транслитерации кириллица выпадала целиком, и «Кресло»
        # с «Диваном» давали одно имя — вторая карточка затирала первую.
        ("кириллица транслитерируется, а не выбрасывается",
         library.slug("Хенге", "Стол") == "henge-stol"
         and library.slug("Хенге", "Кресло") != library.slug("Хенге", "Диван")),
        ("поиск по отделке", find("ashwood") == ["porada-infinity"]),
        ("поиск по артикулу отделки", find("AE") == ["barovier-toso-aurora"]),
        ("поиск по аннотации", find("массива") == ["porada-infinity"]),
        ("несколько слов — нужны все", find("porada стол") == ["porada-infinity"]
         and find("porada лампа") == []),
        ("пустой запрос отдаёт всё", len(find("")) == 2),
    ]
    good = 0
    for label, hit in checks:
        good += hit
        print(f"  {OK if hit else BAD} {label}")

    # Страницы библиотеки закрыты паролем, как и всё остальное.
    import importlib
    import os
    os.environ.setdefault("AURRUM_PASSWORD", "проверка")
    os.environ.setdefault("AURRUM_SECRET_KEY", "x" * 32)
    import app as flask_app
    importlib.reload(flask_app)
    client = flask_app.app.test_client()
    closed = all(client.open(path, method=method).status_code in (301, 302, 401)
                 for path, method in (("/library", "GET"),
                                      ("/library/save", "POST"),
                                      ("/library/delete", "POST")))
    good += closed
    checks.append(("x", closed))
    print(f"  {OK if closed else BAD} библиотека закрыта паролем")
    return good, len(checks)


def check_shops() -> tuple[int, int]:
    """Список магазинов: опознаём по домену, а не по форме адреса.

    `/products/` в пути есть у LONGHI, HENGE и MISURA EMME — по адресу
    магазин неотличим от обычного сайта. Ошибка здесь стоит ложной
    тревоги на каждой карточке этих брендов.
    """
    import gallery
    import shopify
    print("\n ИСТОЧНИК ФОТОГРАФИЙ")
    print(" " + "-" * 74)
    cases = [
        ("https://fendicasa.com/products/annabelle-armchair", True),
        ("https://luxurylivinggroup.com/products/bentley-home-embrace-sofa", True),
        ("https://www.longhi.it/en-us/products/bench-pouf/arianna", False),
        ("https://www.henge07.com/products/tables/sisma/", False),
        # Домен сравниваем целиком: хвостом «fendicasa.com» кончается
        # и чужой домен, зарегистрировать который может кто угодно.
        ("https://evil-fendicasa.com/products/x", False),
    ]
    good = 0
    for url, expected in cases:
        got = shopify.is_shop(url)
        hit = got == expected
        good += hit
        print(f"  {OK if hit else BAD} {url.split('//')[-1][:52]:54} "
              f"-> {'магазин' if got else 'разбор страницы'}")

    # Два источника не должны спорить за один бренд.
    overlap = set(gallery._rules()) & set(shopify.SHOPS)
    hit = not overlap
    good += hit
    print(f"  {OK if hit else BAD} магазины и правила галерей не пересекаются"
          + (f": {', '.join(sorted(overlap))}" if overlap else ""))
    return good, len(cases) + 1


def check_gallery_live() -> tuple[int, int]:
    """Эталонный товар каждого бренда: столько ли снимков отдаёт правило.

    Это единственная проверка, которая ловит переверстку сайта. Дорогая —
    запрос Firecrawl на бренд, — поэтому идёт по ключу `--galleries`,
    а не в каждом прогоне.
    """
    import gallery
    print("\n ГАЛЕРЕИ НА ЖИВЫХ СТРАНИЦАХ")
    print(" " + "-" * 74)
    fc = pl._client()
    good = 0
    rules = gallery._rules()
    for domain, rule in rules.items():
        ref = rule.get("эталон") or {}
        url, want = ref.get("url"), ref.get("фото")
        try:
            _, _, html = pl._scrape(fc, url)
            got = len(gallery.photos(html, url) or [])
        except Exception as exc:  # noqa: BLE001
            print(f"  {BAD} {domain:16} ошибка: {exc}")
            continue
        hit = got == want
        good += hit
        print(f"  {OK if hit else BAD} {domain:16} снимков {got}, ждали {want}")
    return good, len(rules)


def check_login() -> tuple[int, int]:
    """Вход не должен падать на нелатинском пароле.

    Поймано на проде: hmac.compare_digest со строками требует чистого
    ASCII. Достаточно набрать пароль с русской раскладкой — и вместо
    «неверный пароль» приходил 500.
    """
    import os
    os.environ.setdefault("AURRUM_PASSWORD", "проверка-пароля")
    os.environ.setdefault("AURRUM_SECRET_KEY", "x" * 32)
    import importlib
    import app as flask_app
    importlib.reload(flask_app)

    print("\n ВХОД ПО ПАРОЛЮ")
    print(" " + "-" * 74)
    client = flask_app.app.test_client()
    right = os.environ["AURRUM_PASSWORD"]
    cases = [
        ("неверный", 401, "кириллица"),
        ("wrong", 401, "латиница"),
        ("пароль с пробелом и ё", 401, "кириллица с пробелами"),
        ("", 401, "пустой"),
        (right, 302, "верный"),
    ]
    good = 0
    for password, expected, note in cases:
        got = client.post("/login", data={"password": password}).status_code
        hit = got == expected
        good += hit
        print(f"  {OK if hit else BAD} {note:24} -> {got} (ждали {expected})")
    return good, len(cases)


def check_gallery_rules() -> tuple[int, int]:
    """Правила галерей: разбираются ли селекторы и на месте ли эталоны.

    Здесь только форма правила. Саму выборку проверяет `--galleries`:
    он ходит на эталонный товар и сверяет число снимков с записанным.
    Отдельным ключом, потому что это запрос Firecrawl на каждый бренд.
    """
    import gallery
    from bs4 import BeautifulSoup
    print("\n ПРАВИЛА ГАЛЕРЕЙ")
    print(" " + "-" * 74)
    rules = gallery._rules()
    good = 0
    soup = BeautifulSoup("<div></div>", "html.parser")
    for domain, rule in rules.items():
        selectors = rule.get("selectors") or []
        ref = rule.get("эталон") or {}
        try:
            for sel in selectors:
                soup.select(sel)          # синтаксис селектора
            valid = bool(selectors) and bool(ref.get("url")) and bool(ref.get("фото"))
        except Exception:
            valid = False
        good += valid
        print(f"  {OK if valid else BAD} {domain:16} селекторов {len(selectors)}, "
              f"эталон {ref.get('фото', '—')} фото")
    return good, len(rules)


def check_photos() -> tuple[int, int]:
    """Отбор фотографий изделия среди всей галереи страницы.

    У PORADA страница отдавала 72 снимка, к изделию относился 21.
    У VENICEM среди восьми были образцы металла и три других светильника
    той же серии — короткое «circle» захватило бы и их.
    """
    print("\n ОТБОР ФОТОГРАФИЙ")
    print(" " + "-" * 74)
    cases = [
        (["https://x/infinity-01-Infinity.jpg", "https://x/podi-cover.jpg",
          "https://x/02-Infinity.jpg"], "INFINITY", 2, "чужие изделия раздела"),
        (["https://x/circle-floor-1.jpg", "https://x/circle_ceiling_1.jpg",
          "https://x/circle_table_3.jpg", "https://x/materials_metals_M11-1.jpg"],
         "Circle Floor", 1, "другие светильники серии и образцы металла"),
        (["https://x/mere.jpg?width=1946", "https://x/mere.jpg?width=54",
          "https://x/mere-detail.jpg?width=1946"], "Mere", 2, "тот же файл разного размера"),
        (["https://x/a.jpg", "https://x/b.jpg"], "Zeta", 2,
         "по названию не нашлось — отдаём всё"),
        (["https://x/a.jpg", "https://x/logo.png"], "Pin", 1,
         "короткое название не фильтр, логотип отсеян"),
        # Служебные слова сравниваются целиком: подстрока отбрасывала
        # «iconic-collection» из-за «icon», а это частое слово в каталогах.
        (["https://x/iconic-collection/sofa-01.jpg", "https://x/spinello-table.jpg",
          "https://x/avatar-lounge-chair.jpg"], "", 3, "слова внутри других слов"),
        (["https://x/icons/arrow.png", "https://x/img/logo.png",
          "https://x/real-photo.jpg"], "", 1, "каталог icons и файл logo целиком"),
    ]
    good = 0
    for urls, model, want, note in cases:
        got = pl._clean_photos(urls, model)
        hit = len(got) == want
        good += hit
        print(f"  {OK if hit else BAD} {model:14} {len(urls)} -> {len(got):2} (ждали {want})  {note}")
    return good, len(cases)


def check_pricing() -> tuple[int, int]:
    """Цепочка расчёта — против чисел рабочей книги.

    Сверяем СУММУ (AD) и сумму со сборкой (AE): это то, что считается
    формулой. Цену клиенту не сверяем — в книге она округлена по-разному
    (вниз, вверх, до копейки), то есть ставится решением менеджера.
    """
    import pricing
    print("\n РАСЧЁТ ЦЕНЫ ПРОТИВ КНИГИ")
    print(" " + "-" * 74)
    cases = [
        ("LONGHI ARIANA",   6886.0,  0.5, 0.45, 0.00, 1.00,  5752.2,  5752.2),
        ("FLOU BUTTERFLY",  4389.0,  1.0, 0.50, 0.12, 1.00,  4127.8,  4127.8),
        ("VENICEM CIRCLE",  2550.0,  0.2, 0.50, 0.00, 0.95,  2085.0,  2194.7),
        ("TRUSSARDI VIBES", 16020.0, 6.8, 0.50, 0.00, 0.95, 14814.0, 15593.7),
        ("TRUSSARDI COMFY", 5000.0,  0.3, 0.50, 0.00, 1.00,  3850.0,  3850.0),
    ]
    good = 0
    for name, t, vol, disc, markup, asm, want_total, want_assembly in cases:
        r = pricing.line(t, vol, factory_discount=disc, dealer_markup=markup, assembly=asm)
        hit = abs(r.total - want_total) < 0.6 and abs(r.with_assembly - want_assembly) < 0.6
        good += hit
        print(f"  {OK if hit else BAD} {name:18} СУММА {r.total:9.1f} "
              f"со сборкой {r.with_assembly:9.1f}   в книге {want_total:8.1f} / {want_assembly:8.1f}")
    return good, len(cases)


def check_overrides() -> tuple[int, int]:
    """Ручные значения позиции: цена и SWIFT.

    Ровно две ручки, и обе — из формы: цена клиенту в книге округлена
    менеджером (5752,2 -> 5750), SWIFT стоит числом в каждой строке.
    Пустое значение возвращает расчёт.
    """
    import pricing
    print("\n РУЧНЫЕ ЗНАЧЕНИЯ ПОЗИЦИИ")
    print(" " + "-" * 74)
    base = {"list_price": 2550, "volume_m3": 0.2, "qty": 1, "assembly": 0.95}
    plain = pricing.project([base])["lines"][0]
    manual = pricing.project([{**base, "price": 2150, "swift": 0}])["lines"][0]
    cleared = pricing.project([{**base, "price": "", "swift": ""}])["lines"][0]
    checks = [
        ("расчётная цена 2190 — предложение", plain["price"] == 2190),
        ("цена руками 2150 заменяет расчёт", manual["price"] == 2150.0),
        ("сумма считается от ручной цены", manual["sum"] == 2150.0),
        ("SWIFT позиции 0 выигрывает у ставки", manual["swift"] == 0.0),
        # Сравниваем при одинаковом SWIFT: уровни законно зависят от него
        # через СУММУ, а вот ручная цена менять их не должна.
        ("уровни оплаты от ручной цены не зависят",
         manual["levels"] == pricing.project([{**base, "swift": 0}])["lines"][0]["levels"]),
        ("пустое значение возвращает расчёт",
         cleared["price"] == 2190 and cleared["swift"] == 200.0),
    ]
    # Закуп первичен: менеджер знает его из инвойса, а цена прайса
    # выводится обратной арифметикой той же цепочки.
    by_purchase = pricing.project([{"purchase": 3250, "volume_m3": 0.4, "qty": 1,
                                    "factory_discount": 0.5, "dealer_markup": 0,
                                    "assembly": 1}])["lines"][0]
    by_list = pricing.project([{"list_price": 6500, "volume_m3": 0.4, "qty": 1,
                                "factory_discount": 0.5, "dealer_markup": 0,
                                "assembly": 1}])["lines"][0]
    base2 = {"purchase": 3250, "volume_m3": 0.4, "qty": 1,
             "dealer_markup": 0, "assembly": 1}
    checks += [
        ("рентаб процентом: 20 % от закупа",
         pricing.project([{**base2, "margin_pct": 20}])["lines"][0]["margin"] == 650.0),
        ("рентаб евро выигрывает у процента",
         pricing.project([{**base2, "margin_pct": 20, "margin_eur": 1500}])["lines"][0]["margin"] == 1500.0),
        ("нулевой процент законен — «без рентаба»",
         pricing.project([{**base2, "margin_pct": 0}])["lines"][0]["margin"] == 0.0),
        # База транша — цена со скидкой W: при ручном закупе 3250 и
        # наценке 0 это те же 3250, а не выведенный прайс 6500.
        ("транш процентом и евро",
         pricing.project([{**base2, "transfer_pct": 2}])["lines"][0]["transfer"] == 65.0
         and pricing.project([{**base2, "transfer_eur": 50}])["lines"][0]["transfer"] == 50.0),
        ("транспорт ставкой за м³ и евро целиком",
         pricing.project([{**base2, "freight_rate": 300}])["lines"][0]["freight"] == 120.0
         and pricing.project([{**base2, "freight_eur": 1000}])["lines"][0]["freight"] == 1000.0),
        ("закуп руками 3250 -> прайс выведен 6500",
         by_purchase["list_price"] == 6500.0 and by_purchase["purchase"] == 3250.0),
        ("обе дороги дают одну цепочку",
         by_purchase["sum"] == by_list["sum"] and by_purchase["margin"] == by_list["margin"]),
    ]
    good = 0
    for label, hit in checks:
        good += hit
        print(f"  {OK if hit else BAD} {label}")
    return good, len(checks)


def check_final_block() -> tuple[int, int]:
    """Итог компреда — против спецификации 2867.

    В книге два ручных подгона до круглого числа: «+4» в сборке и «−75»
    в скидке. Наши формулы их не содержат — сверяем тождеством: наши
    числа с приложенными подгонами обязаны дать книжные до копейки.
    """
    import pricing
    print("\n ИТОГ КОМПРЕДА ПРОТИВ КНИГИ 2867")
    print(" " + "-" * 74)
    f = pricing.final_block(60370, {"assembly_pct": 5, "personal_pct": 30})
    checks = [
        ("сборка 5 %: 3018.5 (в книге 3022.5 = +4)", f["сборка"] == 3018.5),
        ("всего + подгон 4 = книжные 63392.5", f["всего"] + 4 == 63392.5),
        ("(всего+4)×0.7 − 75 = книжный под-итог 44299.75",
         round((f["всего"] + 4) * 0.7 - 75, 2) == 44299.75),
        ("доп.скидка 4300 доводит до книжных 39999.75",
         round((f["всего"] + 4) * 0.7 - 75 - 4300, 2) == 39999.75),
    ]
    good = 0
    for label, hit in checks:
        good += hit
        print(f"  {OK if hit else BAD} {label}")

    # Блок в выгрузке: формулы ссылаются друг на друга, а не на числа.
    import io
    import book_export
    import openpyxl
    content = book_export.build(
        [{"brand": "X", "model": "Y", "qty": 1, "list_price": 100}],
        final={"assembly_pct": 5, "personal_pct": 30})
    ws = openpyxl.load_workbook(io.BytesIO(content)).active
    labels = [ws.cell(r, 3).value for r in range(14, 28)]
    formulas = [str(ws.cell(r, 6).value) for r in range(14, 28)]
    hit = ("ИТОГО К ОПЛАТЕ, Евро" in labels
           and any(str(v).startswith("=-F") for v in formulas))
    good += hit
    checks.append(("x", hit))
    print(f"  {OK if hit else BAD} блок уходит в файл формулами, включая скидку")
    return good, len(checks)


def check_position_defaults() -> tuple[int, int]:
    """Не заданное поле — число формы, очищенное — ноль.

    Разница дорогая: у нетронутой позиции скидка фабрики обычные 50 %,
    а у очищенной руками её нет вовсе. Если перепутать, цена вырастет
    вдвое или упадёт вдвое — молча, без единого предупреждения.
    """
    import pricing
    print("\n НАЧАЛЬНЫЕ ЧИСЛА ПОЗИЦИИ")
    print(" " + "-" * 74)
    d = pricing.DEFAULT_POSITION
    # Цена 10 000, объём 0: закуп = 10000 x (1 - скидка).
    cases = [
        ("поле не задано -> скидка формы", dict(factory_discount=None),
         10_000 * (1 - d["factory_discount"])),
        ("поле очищено -> без скидки",     dict(factory_discount=""), 10_000.0),
        ("задано число -> оно и берётся",  dict(factory_discount=0.45), 5_500.0),
    ]
    good = 0
    for label, kw, want in cases:
        got = pricing.line(10_000, 0, **kw).purchase
        hit = abs(got - want) < 0.01
        good += hit
        print(f"  {OK if hit else BAD} {label:34} закуп {got:9.1f}, ждали {want:9.1f}")

    # Числа формы: строки 16, 18, 20, 22, 23 рабочего файла.
    hit = (d["factory_discount"], d["dealer_markup"], d["assembly"]) == (0.5, 0.0, 1.0)
    good += hit
    print(f"  {OK if hit else BAD} значения по умолчанию совпадают с формой: "
          f"скидка {d['factory_discount']}, наценка {d['dealer_markup']}, "
          f"сборка {d['assembly']}")
    return good, len(cases) + 1


def check_columns() -> tuple[int, int]:
    """Колонки книги ищутся по подписям, а не по буквам.

    Из-за жёстких букв форма коммерческого предложения читалась как
    спецификация, и в графу «Цена» попадала ширина позиции. Ошибка тихая:
    документ выглядит правильным.
    """
    import spec_parser
    print("\n КОЛОНКИ ПО ПОДПИСЯМ ЗАГОЛОВКА")
    print(" " + "-" * 74)
    headers = [
        ("форма КП",
         {"A": "№", "B": "Производитель", "C": "Описание", "D": "К-во",
          "E": "Цена, Евро", "F": "Сумма, Евро", "G": "СПЕЦ. ЦЕНА, Евро",
          "H": "СПЕЦ. СУММА, Евро"},
         {"n": "A", "brand": "B", "desc": "C", "qty": "D", "price": "E", "total": "F"}),
        ("спецификация",
         {"A": "№", "C": "Производитель", "E": "Описание", "I": "К-во",
          "J": "Цена, Евро", "M": "Сумма, Евро"},
         {"n": "A", "brand": "C", "desc": "E", "qty": "I", "price": "J", "total": "M"}),
        ("спецификация HENGE",
         {"A": "№", "C": "Реф.", "E": "Описание", "I": "К-во",
          "J": "Стоимость, Евро", "M": "Сумма, Евро"},
         {"n": "A", "brand": "C", "desc": "E", "qty": "I", "price": "J", "total": "M"}),
    ]
    good = 0
    for name, cells, expected in headers:
        # ключ вида "B12" -> буква колонки
        got = spec_parser._map_columns(
            lambda ref, c=cells: c.get(ref.rstrip("0123456789")), 12)
        hit = got == expected
        good += hit
        print(f"  {OK if hit else BAD} {name:22} {got}")
        if not hit:
            print(f"      ожидали {expected}")
    return good, len(headers)


def check_book_row() -> tuple[int, int]:
    """Строка для книги: формулы должны совпадать с рабочим файлом.

    Эталон снят с позиции 16 файла 0000-Offer-…-AUR-FORM.xlsx. Если
    разметка книги поедет, проверка это покажет, а не менеджер потом
    в готовом компреде.
    """
    import book_row
    print("\n СТРОКА ДЛЯ КНИГИ")
    print(" " + "-" * 74)
    R = 16
    cells = book_row.visible_row({}, R) + book_row.pricing_row({}, R)
    cols = list("ABCDEFGHIJKLMNOPQRS") + ["T", "U", "V", "W", "X", "Y", "Z",
                                          "AA", "AB", "AC", "AD", "AE", "AF",
                                          "AG", "AH", "AI"]
    got = dict(zip(cols, (str(c) for c in cells)))
    expected = {
        "F": "=D16*E16", "H": "=D16*G16", "R": "=U16", "S": "=R16*D16",
        "U": "=ROUNDUP(J16*K16*L16*1.5/1000000,1)",
        "W": "=T16-T16*V16", "Y": "=W16+W16*X16", "Z": "=Y16*($Z$1+0)/100",
        "AA": "=W16*$AA$1/100", "AC": "=U16*$AC$1",
        "AD": "=Y16+Z16+AA16+AB16+AC16", "AE": "=AD16/1",
        "AF": "=AE16/(1-$AF$1/100)", "AG": "=AF16/(1-$AG$1/100)",
        "AH": "=AG16/(1-$AH$1/100)", "AI": "=AH16/(1-$AI$1/100)",
    }
    good = 0
    for col, want in expected.items():
        hit = got.get(col) == want
        good += hit
        print(f"  {OK if hit else BAD} {col:3} {want}")
        if not hit:
            print(f"      у нас: {got.get(col)}")
    # Многострочное описание обязано пережить вставку
    multi = book_row.build({"description": "A\nB"}, R)
    quoted = '"A\nB"' in multi
    good += quoted
    print(f"  {OK if quoted else BAD} описание в кавычках — перевод строки не рвёт строку")
    return good, len(expected) + 1


def check_docs_list() -> tuple[int, int]:
    """Список документов: бумаги сайта в него не попадают.

    На каждой странице магазина висит заявление о доступности, у BAROVIER —
    политика информирования. Инструкция по сборке при этом остаётся:
    она про изделие, просто не техлист.
    """
    print("\n СПИСОК ДОКУМЕНТОВ")
    print(" " + "-" * 74)
    links = [
        "https://cdn.shopify.com/s/files/Accessibility_Statement_LANG_EN.pdf?v=176",
        "https://x/cdn/shop/files/VIBES_bed_GUEST.pdf?v=674412",
        "https://venicem.com/Venicem_Assembly_Instructions_Circle_Floor.pdf",
        "https://venicem.com/Venicem_product_fact_sheet_circle-floor.pdf",
        "https://barovier.com/files/policy-whistleblowing-barovier-toso_0.pdf",
        "https://x/page.html",
    ]
    docs = pl._docs_from(links)
    names = [d.rsplit("/", 1)[-1].split("?")[0] for d in docs]
    checks = [
        ("заявление о доступности скрыто",
         not any("Accessibility" in n for n in names)),
        ("политика информирования скрыта",
         not any("whistleblowing" in n for n in names)),
        ("инструкция по сборке осталась",
         any("Assembly" in n for n in names)),
        ("техлист остался", any("fact_sheet" in n for n in names)),
        ("не-PDF отсеян", all(n.endswith(".pdf") for n in names)),
    ]
    good = 0
    for note, hit in checks:
        good += hit
        print(f"  {OK if hit else BAD} {note}")
    return good, len(checks)


def check_spec_pdf() -> tuple[int, int]:
    """Отбор техлиста среди прочих PDF страницы."""
    print("\n ОТБОР ТЕХЛИСТА")
    print(" " + "-" * 74)
    cases = [
        (["https://x/cdn/shop/files/VIBES_bed_GUEST.pdf?v=674412",
          "https://cdn.shopify.com/s/files/Accessibility_Statement_EN.pdf?v=176"],
         "VIBES", "адрес с хвостом версии, рядом заявление о доступности"),
        (["https://barovier.com/files/policy-whistleblowing-barovier-toso_0.pdf"],
         "", "единственный PDF — политика, техлиста нет"),
        (["https://venicem.com/Venicem_Assembly_Instructions_Pinocchio.pdf",
          "https://venicem.com/Venicem_product_fact_sheet_pinocchio-floor.pdf"],
         "fact_sheet", "инструкция сборки рядом с техлистом"),
        ([], "", "документов нет"),
    ]
    good = 0
    for urls, expected, note in cases:
        got = pl._pick_spec_pdf(urls)
        hit = (expected in got) if expected else (got == "")
        good += hit
        print(f"  {OK if hit else BAD} {note:56} -> {(got.rsplit('/', 1)[-1] or '—')[:26]}")
    return good, len(cases)


def check_url_types() -> tuple[int, int]:
    print("\n ТИП ПО РАЗДЕЛУ САЙТА")
    print(" " + "-" * 74)
    good = 0
    for url, expected in URL_TYPE_CASES:
        got = pl.type_from_url(url)
        hit = got == expected
        good += hit
        short = url.split("//")[-1][:52]
        print(f"  {OK if hit else BAD} {short:54} -> {got or '—'}")
        if not hit:
            print(f"      ожидали {expected or '—'}")
    return good, len(URL_TYPE_CASES)


def check_type_norm() -> tuple[int, int]:
    print("\n ТИП ИЗ ИЗВЛЕЧЕНИЯ")
    print(" " + "-" * 74)
    good = 0
    for raw, expected in TYPE_NORM_CASES:
        got, _ = pl.normalize_type(raw)
        hit = got == expected
        good += hit
        print(f"  {OK if hit else BAD} {(raw or '—'):20} -> {got or '—'}")
        if not hit:
            print(f"      ожидали {expected or '—'}")
    return good, len(TYPE_NORM_CASES)


def check_schema() -> tuple[int, int]:
    """Схема извлечения дублирует списки из кода — стережём расхождение.

    Заодно ловим форму, на которой LlamaExtract падает ещё до чтения
    документа: nullable-массив он разворачивает в anyOf и теряет items.
    """
    print("\n СХЕМА ИЗВЛЕЧЕНИЯ (config/extraction_schema.json)")
    print(" " + "-" * 74)
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "config", "extraction_schema.json")
    schema = json.load(open(path, encoding="utf-8"))
    item = schema["properties"]["products"]["items"]["properties"]

    bad: list[str] = []

    def walk(node, where="корень"):
        t = node.get("type")
        if isinstance(t, list) and "array" in t:
            bad.append(f"{where}: массив объявлен nullable — LlamaExtract потеряет items")
        if t == "array" and "items" not in node:
            bad.append(f"{where}: массив без items")
        # Проверено опытом: enum LlamaExtract не соблюдает, а молча отдаёт
        # последнее значение списка. Кровать превращалась в «Ковёр»,
        # ткань — в «Камень». Допустимые значения держим в description.
        if "enum" in node:
            bad.append(f"{where}: enum — LlamaExtract вернёт последнее значение списка")
        for k, v in (node.get("properties") or {}).items():
            walk(v, f"{where}.{k}")
        if isinstance(node.get("items"), dict):
            walk(node["items"], f"{where}[]")

    walk(schema)

    role = item["finishes"]["items"]["properties"]["role_ru"]
    checks = [
        ("все типы перечислены в описании type_ru",
         all(t in item["type_ru"]["description"] for t in pl.TYPES_RU)),
        ("все роли перечислены в описании role_ru",
         all(r in role["description"] for r in pl.ROLES_RU)),
        ("исполнение — строка массива variants", "dims_raw" in
         item["variants"]["items"]["properties"]),
        ("форма схемы принимается LlamaExtract", not bad),
    ]
    for label, hit in checks:
        print(f"  {OK if hit else BAD} {label}")
    for b in bad:
        print(f"      {b}")
    return sum(h for _, h in checks), len(checks)


def _same(expected, actual) -> bool:
    """Сравнение с эталоном: числа сверяем как числа, а не как строки."""
    if expected is None or actual is None:
        return expected == actual
    try:
        return abs(float(expected) - float(actual)) < 0.01
    except (TypeError, ValueError):
        return str(expected).strip().lower() == str(actual).strip().lower()


def show(p, expect: dict | None = None) -> None:
    """Печать карточки; расхождения с ожиданием помечаются."""
    expect = expect or {}

    def line(label, value, key=None, actual=None):
        mark = " "
        if key in expect:
            mark = OK if _same(expect[key], value if actual is None else actual) else BAD
        print(f"   {mark} {label:14} {value}")

    def dims_line():
        keys = ("width_cm", "depth_cm", "height_cm")
        vals = (p.width_cm, p.depth_cm, p.height_cm)
        shown = " / ".join(f"{v:g}" if v else "—" for v in vals)
        checked = [k for k in keys if k in expect]
        mark = " "
        if checked:
            mark = OK if all(_same(expect[k], v) for k, v in zip(keys, vals) if k in expect) else BAD
        print(f"   {mark} {'Д / Г / В':14} {shown}")

    line("производитель", p.brand or "—")
    line("модель", p.model or "—")
    line("тип", p.type_ru or "—", "type_ru")
    line("размеры", p.dims_raw or "—")
    dims_line()
    line("объём", f"{p.volume_m3 or '—'} м³ ({p.volume_source or 'не определён'})")
    line("отделки", ", ".join(
        f"{f.get('role_ru')}: {f.get('material')}" for f in p.finishes) or "—")
    line("примечание", p.tech_note or "—")
    line("фото / доки", f"{len(p.photo_urls)} / {len(p.doc_urls)}")
    for w in p.warnings:
        print(f"   {WARN} {w}")


def check_live() -> None:
    print("\n ЖИВЫЕ ССЫЛКИ (нужен FIRECRAWL_API_KEY)")
    print(" " + "-" * 74)
    for case in LIVE_CASES:
        print(f"\n  {case['url']}")
        print(f"  {case['note']}")
        try:
            p = pl.lookup(case["url"])
        except Exception as exc:  # noqa: BLE001
            print(f"   {BAD} ошибка: {exc}")
            continue
        show(p, case.get("expect"))
        print("   описание для колонки C:")
        for ln in pl.to_excel_description(p).splitlines():
            print(f"      | {ln}")


def main() -> int:
    args = sys.argv[1:]

    if args and args[0].startswith("http"):
        p = pl.lookup(args[0])
        print(f"\n {args[0]}")
        show(p)
        print("\n описание для колонки C:")
        for ln in pl.to_excel_description(p).splitlines():
            print(f"   | {ln}")
        return 0

    d_ok, d_all = check_dims()
    v_ok, v_all = check_volume()
    g_ok, g_all = check_dims_grounding()
    f_ok, f_all = check_spec_pdf()
    d_ok, d_all = check_docs_list()
    b_ok, b_all = check_book_row()
    c_ok, c_all = check_columns()
    p_ok, p_all = check_pricing()
    pd_ok, pd_all = check_position_defaults()
    fb_ok, fb_all = check_final_block()
    ov_ok, ov_all = check_overrides()
    r_ok, r_all = check_gallery_rules()
    l_ok, l_all = check_login()
    ph_ok, ph_all = check_photos()
    t_ok, t_all = check_url_types()
    tn_ok, tn_all = check_type_norm()
    sh_ok, sh_all = check_shops()
    lb_ok, lb_all = check_library()
    hd_ok, hd_all = check_header_roundtrip()
    dl_ok, dl_all = check_download_headers()
    s_ok, s_all = check_schema()

    gl_ok = gl_all = None
    if "--galleries" in args:
        gl_ok, gl_all = check_gallery_live()

    if "--offline" not in args and "--galleries" not in args:
        check_live()

    ok = (d_ok == d_all and v_ok == v_all and s_ok == s_all
          and t_ok == t_all and g_ok == g_all and f_ok == f_all
          and b_ok == b_all and c_ok == c_all and p_ok == p_all
          and ph_ok == ph_all and tn_ok == tn_all and r_ok == r_all
          and sh_ok == sh_all and pd_ok == pd_all and lb_ok == lb_all and hd_ok == hd_all and dl_ok == dl_all and fb_ok == fb_all and ov_ok == ov_all
          and (gl_all is None or gl_ok == gl_all))
    print("\n" + " " + "=" * 74)
    print(f"  Размеры: {d_ok}/{d_all}   Объём: {v_ok}/{v_all}   "
          f"Сверка: {g_ok}/{g_all}   Техлист: {f_ok}/{f_all}   "
          f"Строка: {b_ok}/{b_all}   Колонки: {c_ok}/{c_all}   Расчёт: {p_ok}/{p_all}   "
          f"Начальные числа: {pd_ok}/{pd_all}   Итог: {fb_ok}/{fb_all}   Ручные: {ov_ok}/{ov_all}   Шапка: {hd_ok}/{hd_all}   "
          f"Выгрузка: {dl_ok}/{dl_all}\n"
          f"  Фото: {ph_ok}/{ph_all}   Правила галерей: {r_ok}/{r_all}   "
          f"Источник фото: {sh_ok}/{sh_all}   Библиотека: {lb_ok}/{lb_all}   "
          f"Тип по адресу: {t_ok}/{t_all}   Тип из извлечения: {tn_ok}/{tn_all}   "
          f"Схема: {s_ok}/{s_all}"
          + (f"\n  Галереи на живых страницах: {gl_ok}/{gl_all}"
             if gl_all is not None else ""))
    if ok:
        print("  Разбор совпадает с книгой, схема согласована с кодом.")
    else:
        print("  ЕСТЬ РАСХОЖДЕНИЯ — смотрите строки с ✗ выше.")
    print(" " + "=" * 74)
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
