# -*- coding: utf-8 -*-
"""
Проект -> готовый файл Excel в разметке рабочей формы.

Закрывает последний ручной шаг: раньше строки переносились через буфер
по одной, а фотографии менеджер вставлял руками. Здесь позиции, формулы,
итоги и снимки уже на местах — файл открывают и работают дальше в Excel.

Ставки пишутся в первую строку (`Z1`, `AA1`, `AC1`, `AF1`…`AI1`): формулы
расчёта ссылаются именно на них, и без этих ячеек цена не посчитается.
Заодно они снимаются в файл — открытый через полгода проект покажет те
числа, что были при отправке, а не пересчитается по новым ставкам.

Фотография кладётся одна на позицию, первая из отмеченных: столько же
их в рабочей форме, и ответ не распухает — на serverless он ограничен.
"""

from __future__ import annotations

import io

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import book_row
import pricing
import safe_fetch

# Разметка формы: ставки в первой строке, заголовок таблицы в 13-й.
RATES_ROW = 1
HEADER_ROW = 13
FIRST_ITEM_ROW = HEADER_ROW + 1

# Шапка — ровно там, где её ищет обратный разбор (`spec_parser`), а он
# писан с живой книги `samples/2867_Спецификация_…`:
#
#   A5 «Покупатель»   H5 «Договор»   M5 «Дата»
#   A6  значение      H6  значение   M6  значение
#   A8 «Спецификация №… к Договору …»              N8 дата
#
# Строка 8 дублирует номер и договор нарочно: так сделано в книге, и
# разбор берёт их оттуда первым делом. Без этих ячеек напечатанный
# документ выходил с пустым номером и прочерками в покупателе и дате.
HEAD_LABEL_ROW = 5
HEAD_VALUE_ROW = 6
HEAD_TITLE_ROW = 8

# Ячейки ставок — те же, на которые ссылаются формулы позиции.
RATE_CELLS = {
    "Z": ("margin", "РЕНТАБ, %"),
    "AA": ("transfer", "ТРАНШ, %"),
    "AC": ("freight", "ТРАНСПОРТ, евро за м³"),
    "AF": ("designer", "ДИЗАЙНЕР, %"),
    "AG": ("usno", "УСНО, %"),
    "AH": ("vat", "НДС, %"),
    "AI": ("finserv", "FINSERV, %"),
}

HEADERS = [
    "№", "Производитель", "Описание", "К-во", "Цена, Евро", "Сумма, Евро",
    "СПЕЦ. ЦЕНА, Евро", "СПЕЦ. СУММА, Евро", "Фото / Схема",
    "Д, см", "Г, см", "В, см", "Отделка 1", "Отделка 2", "Отделка 3",
    "Примечание", "Схема", "м3", "м3 всего",
]

# Ширина колонки с фото и высота строки под снимок — в точках Excel.
PHOTO_COLUMN_WIDTH = 24
PHOTO_ROW_HEIGHT = 96
PHOTO_MAX_PX = 170

COLUMN_WIDTHS = {"A": 5, "B": 18, "C": 46, "D": 6, "E": 12, "F": 12,
                 "G": 14, "H": 14, "I": PHOTO_COLUMN_WIDTH,
                 "J": 8, "K": 8, "L": 8, "M": 12, "N": 12, "O": 12,
                 "P": 18, "Q": 10, "R": 8, "S": 8}


def _number(value):
    """Строку с числом пишем числом — иначе Excel не посчитает по ней."""
    if isinstance(value, (int, float)):
        return value
    text = str(value or "").strip().replace(",", ".")
    if not text or text.startswith("="):
        return value
    try:
        number = float(text)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def _photo(url: str, max_px: int = PHOTO_MAX_PX) -> XLImage | None:
    """Скачиваем и ужимаем снимок. Не вышло — молча пропускаем: файл
    без картинки лучше, чем отсутствие файла.

    `max_px` зависит от адресата: в ячейку Excel хватает 170, а печать
    показывает снимок втрое крупнее и из 170 получает мыло — ей 900,
    как и загруженным книгам (`spec_parser._to_data_uri`).
    """
    try:
        raw = safe_fetch.get(url, timeout=30,
                             headers={"User-Agent": "Mozilla/5.0"}).content
        from PIL import Image

        im = Image.open(io.BytesIO(raw))
        im.thumbnail((max_px, max_px), Image.LANCZOS)
        if im.mode != "RGB":
            im = im.convert("RGB")
        buf = io.BytesIO()
        im.save(buf, "JPEG", quality=82, optimize=True)
        buf.seek(0)
        return XLImage(buf)
    except Exception:            # noqa: BLE001 — картинка не критична
        return None


def _header(ws, head: dict) -> None:
    """Покупатель, договор, номер и дата — в ячейки разметки книги."""
    number = str(head.get("number") or "").strip()
    contract = str(head.get("contract") or "").strip()
    buyer = str(head.get("buyer") or "").strip()
    day = str(head.get("date") or "").strip()

    for column, label, value in (("A", "Покупатель", buyer),
                                 ("H", "Договор", contract),
                                 ("M", "Дата", day)):
        ws[f"{column}{HEAD_LABEL_ROW}"] = label
        ws[f"{column}{HEAD_LABEL_ROW}"].font = Font(size=9, color="949598")
        ws[f"{column}{HEAD_VALUE_ROW}"] = value

    # Строку заголовка собираем, только если есть номер: пустое
    # «Спецификация № к Договору» хуже, чем его отсутствие.
    if number:
        title = f"Спецификация №{number}"
        if contract:
            title += f" к Договору {contract}"
        ws[f"A{HEAD_TITLE_ROW}"] = title
        ws[f"A{HEAD_TITLE_ROW}"].font = Font(size=11, bold=True)
        ws[f"N{HEAD_TITLE_ROW}"] = day


def build(positions: list[dict], rates: dict | None = None,
          header: dict | None = None, final: dict | None = None,
          values: bool = False) -> bytes:
    """Позиции проекта -> содержимое файла .xlsx.

    `values=True` — те же ячейки числами вместо формул: печать разбирает
    книгу тут же, а формулы в свежем файле ещё никем не вычислены, и
    разбор увидел бы пустоту. Числа считает тот же `pricing` — правда
    одна, отличается только форма записи.
    """
    r = {**pricing.DEFAULT_RATES, **(rates or {})}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"

    # Ставки и подписи к ним. Без первой строки формулы позиций молчат.
    for column, (key, label) in RATE_CELLS.items():
        ws[f"{column}{RATES_ROW}"] = r[key]
        ws[f"{column}{RATES_ROW + 1}"] = label
        ws[f"{column}{RATES_ROW + 1}"].font = Font(size=8, color="949598")

    _header(ws, header or {})

    ws["A9"] = "Коммерческое предложение по решению интерьера"
    ws["A9"].font = Font(size=13, bold=True)

    for index, title in enumerate(HEADERS, start=1):
        cell = ws.cell(HEADER_ROW, index, title)
        cell.font = Font(size=9, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")

    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width

    for offset, position in enumerate(positions):
        row = FIRST_ITEM_ROW + offset
        computed = pricing.line(
            position.get("list_price"), position.get("volume_m3"),
            factory_discount=position.get("factory_discount"),
            dealer_markup=position.get("dealer_markup"),
            assembly=position.get("assembly"), rates=r,
        )
        fields = {**position, "number": offset + 1, "swift": r["swift"],
                  # Цена клиенту — предложение расчёта; менеджер правит в файле.
                  "price": position.get("price") or computed.price or ""}

        cells = book_row.visible_row(fields, row) + book_row.pricing_row(fields, row)
        if values:
            qty = max(1, int(pricing._num(position.get("qty"), 1)))
            price = pricing._num(fields.get("price"))
            volume = pricing._num(position.get("volume_m3"))
            cells[5] = round(price * qty, 2) if price else ""   # F  Сумма
            cells[17] = volume or ""              # R  м3
            cells[18] = round(volume * qty, 2) if volume else ""   # S  м3 всего
        for index, value in enumerate(cells, start=1):
            ws.cell(row, index, _number(value))

        ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = PHOTO_ROW_HEIGHT

        photos = position.get("photos") or []
        if photos:
            image = _photo(photos[0], max_px=900 if values else PHOTO_MAX_PX)
            if image is not None:
                ws.add_image(image, f"I{row}")

    # Итоговый блок компреда — формулами, цепочка из спецификации 2867:
    # проценты вписаны в формулы числами, как делает сама книга
    # (=M17*0.05), поэтому файл пересчитывается в Excel без нас.
    last = FIRST_ITEM_ROW + max(0, len(positions)) - 1
    totals_row = last + 2
    if positions:
        f = {**pricing.DEFAULT_FINAL,
             **{k: v for k, v in (final or {}).items() if isinstance(v, (int, float))}}
        t = totals_row
        if values:
            items_sum = sum(
                pricing._num(p.get("price")) * max(1, int(pricing._num(p.get("qty"), 1)))
                for p in positions) or sum(
                pricing.line(p.get("list_price"), p.get("volume_m3"),
                             factory_discount=p.get("factory_discount"),
                             dealer_markup=p.get("dealer_markup"),
                             assembly=p.get("assembly"), rates=r).price
                * max(1, int(pricing._num(p.get("qty"), 1))) for p in positions)
            fb = pricing.final_block(items_sum, f)
            rows = [
                ("Сумма, Евро", round(items_sum, 2), True),
                ("Дополнительные услуги, Евро", fb["услуги"], False),
                ("Доставка по Москве/МО, Евро", fb["доставка"], False),
                ("Сборка/Монтаж, Евро", fb["сборка"], False),
                ("Всего, Евро", fb["всего"], True),
                ("Исключительная Персональная Скидка, Евро", fb["скидка"], False),
                ("Под-Итог, Евро", fb["подытог"], True),
                ("Дополнительная Скидка, Евро", fb["доп_скидка"], False),
                ("ИТОГО К ОПЛАТЕ, Евро", fb["к_оплате"], True),
            ]
        else:
            def part(key, base_cell, sign=""):
                # Ненулевое евро выигрывает у процента — как в pricing.
                if f[f"{key}_eur"]:
                    value = round(abs(f[f"{key}_eur"]), 2)
                    return -value if sign == "-" else value
                return f"={sign}{base_cell}*{f[f'{key}_pct'] / 100}"

            rows = [
                ("Сумма, Евро", f"=SUM(F{FIRST_ITEM_ROW}:F{last})", True),
                ("Дополнительные услуги, Евро", part("services", f"F{t}"), False),
                ("Доставка по Москве/МО, Евро", part("delivery", f"F{t}"), False),
                ("Сборка/Монтаж, Евро", part("assembly", f"F{t}"), False),
                ("Всего, Евро", f"=F{t}+F{t + 1}+F{t + 2}+F{t + 3}", True),
                ("Исключительная Персональная Скидка, Евро",
                 part("personal", f"F{t + 4}", sign="-"), False),
                ("Под-Итог, Евро", f"=F{t + 4}+F{t + 5}", True),
                ("Дополнительная Скидка, Евро", -abs(f["extra_eur"]) or 0, False),
                ("ИТОГО К ОПЛАТЕ, Евро", f"=F{t + 6}+F{t + 7}", True),
            ]
        for offset, (label, value, strong) in enumerate(rows):
            row = t + offset
            ws.cell(row, 3, label).font = Font(bold=strong)
            cell = ws.cell(row, 6, _number(value))
            cell.font = Font(bold=strong)
        ws.cell(t, 19, f"=SUM(S{FIRST_ITEM_ROW}:S{last})")

    # Служебные колонки прячем, как в рабочей форме.
    for column in ("G", "R"):
        ws.column_dimensions[column].hidden = True
    for index in range(20, 36):
        ws.column_dimensions[get_column_letter(index)].hidden = True

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
