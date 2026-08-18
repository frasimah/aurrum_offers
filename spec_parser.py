# -*- coding: utf-8 -*-
"""
Разбор Excel-спецификации AURRUM в структуру для печатного документа.

Важное про приватность данных: в исходной книге есть внутренние колонки
(закупка, пошлина, логистика, лестница наценок) и внутренние строки
(рабочая скидка, комиссии, оплата напрямую на фабрику). В печать они
не идут. Ориентир — флаги hidden самой книги: то, что скрыто в Excel,
скрыто и здесь.
"""

from __future__ import annotations

import base64
import io
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

NS_XDR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
NS_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
NS_R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

# Запасная разметка — шаблон «Спецификация». Берётся, только если подписи
# в строке заголовка прочитать не удалось.
COL_N, COL_BRAND, COL_DESC = "A", "C", "E"
COL_QTY, COL_PRICE, COL_SUM = "I", "J", "M"

# Колонки ищем по подписям заголовка, а не по буквам: у книги «Спецификация»
# описание лежит в E, а у формы коммерческого предложения — в C. С жёсткими
# буквами форма читалась как спецификация, и в графу «Цена» попадала ширина
# позиции. Ошибка тихая — документ выглядит правильным.
_HEADER_KEYS = {
    "n": ("№",),
    "brand": ("производител", "реф"),
    "desc": ("описание",),
    "qty": ("к-во", "кол-во", "количество"),
    "price": ("цена", "стоимость"),
    "total": ("сумма",),
}

# «СПЕЦ. ЦЕНА» и «СПЕЦ. СУММА» — внутренние колонки для торга. В документ
# они не идут и не должны перехватывать поиск цены.
_HEADER_SKIP = ("спец",)

# Максимум строк, которые просматриваем в поисках позиций и итогов
SCAN_LIMIT = 200


@dataclass
class Item:
    n: str = ""
    brand: str = ""
    title: str = ""
    body: str = ""
    qty: str = ""
    price: float | None = None
    total: float | None = None
    photo: str | None = None


@dataclass
class Spec:
    number: str = ""
    contract: str = ""
    date: str = ""
    buyer: str = ""
    section_title: str = "Общая информация"
    origin: str = ""
    badge: str = ""
    items: list[Item] = field(default_factory=list)
    totals: list[dict] = field(default_factory=list)
    terms: list[str] = field(default_factory=list)
    terms_note: str = ""
    warnings: list[str] = field(default_factory=list)


# --- Условия. В книге они капсом и почти всегда одинаковые, поэтому
#     нормализуем регистр и держим как редактируемый текст документа.
DEFAULT_TERMS = [
    "Указаны цены нетто при условии оплаты наличными",
    "Оплата в рублях по внутреннему курсу на момент оплаты",
    "Предоплата 70%, баланс 30% по готовности к отгрузке с фабрики-производителя",
    "Срок поставки в Москву — 20 недель (не включая август)",
    "Фото приведены из каталога и служат лишь для понимания модели изделия",
    "Предлагаемый предмет может отличаться по комплектации, отделке, "
    "наполнению от приведённого фото",
    "Подробное описание предлагаемого предмета со всеми опциями приведено "
    "в колонке «Описание»",
    "Любые изменения, дополнения в проект подлежат согласованию с фабрикой "
    "и уточнению стоимости",
]
DEFAULT_TERMS_NOTE = (
    "Ручной подъём на этаж и/или подъём альпинистами/краном оплачиваются "
    "отдельно по согласованию при необходимости"
)


def _fmt_date(v) -> str:
    if isinstance(v, (datetime, date)):
        return v.strftime("%d.%m.%Y")
    return str(v).strip() if v is not None else ""


# Имена собственные, которые нельзя терять при снятии капса.
_PROPER = {"москва": "Москва", "москву": "Москву", "москве": "Москве"}

# Точечные исправления текста условий. Список намеренно короткий и обозримый,
# правится руками — авто-исправление грамматики регулярками по документу,
# который идёт под подпись, делать не стоит.
_FIXUPS = [
    (r"\bна основание\b", "на основании"),
    # В книге пишут без «ё» — возвращаем там, где это однозначно.
    (r"\bрасчет\b", "расчёт"),
    (r"\bпроизведен\b", "произведён"),
    (r"\bприведенн(\w*)", r"приведённ\1"),
    (r"\bподъем\b", "подъём"),
]


def _normalize_term(s: str) -> str:
    """КАПС -> обычный регистр + типографика. Некапсовые строки почти не трогаем."""
    s = re.sub(r"\s{2,}", " ", s.strip())
    letters = [c for c in s if c.isalpha()]
    if letters and all(c.isupper() for c in letters):
        s = s.lower()

    s = re.sub(r"\b(москва|москву|москве)\b", lambda m: _PROPER[m.group(1)], s)
    for pat, repl in _FIXUPS:
        s = re.sub(pat, repl, s)

    s = re.sub(r"(\d{2})-(\d{2})-(\d{4})", r"\1.\2.\3", s)   # 06-03-2026 -> 06.03.2026
    s = re.sub(r'"([^"]*)"', r"«\1»", s)                      # "..." -> «...»
    s = re.sub(r"«(\w)", lambda m: "«" + m.group(1).upper(), s)
    s = re.sub(r"(?<=\s)-(?=\s)", "—", s)                     # дефис-разделитель -> тире

    return s[:1].upper() + s[1:] if s else s


# В печати колонка фото — 40 мм, поэтому 900 px хватает с большим запасом.
# Ужимаем осознанно: документ отдаётся одним ответом, а на serverless он
# ограничен по размеру.
MAX_PHOTO_PX = 900
JPEG_QUALITY = 82


def _to_data_uri(raw: bytes, ext: str) -> str:
    """Картинка -> data-URI. Крупные ужимаем, чтобы документ не распухал."""
    mime = "image/png" if ext.lower() == ".png" else "image/jpeg"
    try:
        from PIL import Image

        im = Image.open(io.BytesIO(raw))
        if max(im.size) > MAX_PHOTO_PX:
            im.thumbnail((MAX_PHOTO_PX, MAX_PHOTO_PX), Image.LANCZOS)
        if im.mode in ("RGBA", "LA", "P"):
            im = im.convert("RGBA")
            flat = Image.new("RGB", im.size, (255, 255, 255))
            flat.paste(im, mask=im.split()[-1])  # прозрачность -> белый фон
            im = flat
        elif im.mode != "RGB":
            im = im.convert("RGB")
        buf = io.BytesIO()
        im.save(buf, "JPEG", quality=JPEG_QUALITY, optimize=True)
        raw, mime = buf.getvalue(), "image/jpeg"
    except Exception:
        pass  # Pillow нет или формат экзотический — отдаём как есть

    return f"data:{mime};base64," + base64.b64encode(raw).decode("ascii")


def _extract_photos(data: bytes) -> dict[int, str]:
    """Картинки -> {номер строки листа: data-URI}. Привязка берётся из drawing.

    Ничего не пишем на диск: документ должен быть самодостаточным, а на
    serverless-хостинге постоянного диска между запросами и нет.
    """
    photos: dict[int, str] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            names = z.namelist()
            drawings = [n for n in names if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
            for d in drawings:
                rels_path = f"xl/drawings/_rels/{os.path.basename(d)}.rels"
                rels = {}
                if rels_path in names:
                    for r in ET.fromstring(z.read(rels_path)):
                        # Excel пишет путь относительным (../media/…),
                        # openpyxl — абсолютным (/xl/media/…). Обе формы
                        # приводим к имени внутри архива.
                        rels[r.get("Id")] = (r.get("Target", "")
                                             .replace("../", "xl/").lstrip("/"))
                for anchor in ET.fromstring(z.read(d)):
                    frm = anchor.find(f"{NS_XDR}from")
                    blip = anchor.find(f".//{NS_A}blip")
                    if frm is None or blip is None:
                        continue
                    row = int(frm.find(f"{NS_XDR}row").text) + 1  # 0-индекс -> 1-индекс
                    target = rels.get(blip.get(f"{NS_R}embed"))
                    if not target or target not in names:
                        continue
                    ext = os.path.splitext(target)[1] or ".png"
                    photos[row] = _to_data_uri(z.read(target), ext)
    except (zipfile.BadZipFile, ET.ParseError, KeyError):
        pass  # книга без картинок — не повод падать
    return photos


def _map_columns(g, header_row: int) -> dict[str, str]:
    """Буквы колонок по подписям строки заголовка.

    Первое совпадение слева выигрывает, поэтому «Цена, Евро» находится
    раньше, чем «СПЕЦ. ЦЕНА», а та вдобавок отсеивается по слову «спец».
    """
    found: dict[str, str] = {}
    for c in range(1, 40):
        label = g(f"{get_column_letter(c)}{header_row}")
        if not isinstance(label, str):
            continue
        low = " ".join(label.lower().split())
        if not low or any(skip in low for skip in _HEADER_SKIP):
            continue
        for key, needles in _HEADER_KEYS.items():
            if key not in found and any(n in low for n in needles):
                found[key] = get_column_letter(c)
    return found


def parse(data: bytes) -> Spec:
    """Книга приходит байтами: временные файлы не нужны ни локально,
    ни на serverless, где диск между запросами не сохраняется."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]
    g = lambda c: ws[c].value  # noqa: E731

    spec = Spec()
    photos = _extract_photos(data)
    hidden_rows = {k for k, d in ws.row_dimensions.items() if d.hidden}

    # --- Шапка: "Спецификация №2867/3 к Договору 2867"
    for r in range(1, 12):
        v = g(f"A{r}")
        if isinstance(v, str) and "спецификация" in v.lower():
            m = re.search(r"№\s*([\w/\-]+)", v)
            if m:
                spec.number = m.group(1)
            m = re.search(r"договор\w*\s+([\w/\-]+)", v, re.I)
            if m:
                spec.contract = m.group(1)
            spec.date = _fmt_date(g(f"N{r}"))
            break

    # Покупатель / Договор / Дата — блок подписан в одной строке, значения ниже
    for r in range(1, 12):
        if isinstance(g(f"A{r}"), str) and g(f"A{r}").strip().lower() == "покупатель":
            spec.buyer = str(g(f"A{r+1}") or "").strip()
            if not spec.contract:
                spec.contract = str(g(f"H{r+1}") or "").strip()
            if not spec.date:
                spec.date = _fmt_date(g(f"M{r+1}"))
            break

    # Заголовок таблицы позиций: в колонке A стоит "№". От него отсчитываем
    # всё остальное — разбор не ломается, если шапка книги сдвинулась.
    header_row = None
    for r in range(1, SCAN_LIMIT):
        if str(g(f"A{r}") or "").strip() == "№":
            header_row = r
            break
    if header_row is None:
        header_row = 12  # как в текущем шаблоне книги
        spec.warnings.append(
            "Не найден заголовок таблицы (ячейка со знаком «№» в колонке A) — "
            "разбираю по стандартной разметке."
        )

    cols = _map_columns(g, header_row)
    col_n = cols.get("n", COL_N)
    col_brand = cols.get("brand", COL_BRAND)
    col_desc = cols.get("desc", COL_DESC)
    col_qty = cols.get("qty", COL_QTY)
    col_price = cols.get("price", COL_PRICE)
    col_sum = cols.get("total", COL_SUM)
    col_total_value = col_sum
    if len(cols) < len(_HEADER_KEYS):
        spec.warnings.append(
            "В заголовке таблицы нашлись не все колонки — часть разобрана "
            "по запасной разметке. Проверьте цену и количество."
        )

    # Строка над заголовком: страна происхождения и пометка вроде «НА ЗАКАЗ»
    meta_row = header_row - 1
    origin = g(f"A{meta_row}")
    if isinstance(origin, str) and origin.strip().startswith("("):
        spec.origin = origin.strip()
    badge = g(f"N{meta_row}")
    if isinstance(badge, str) and badge.strip():
        spec.badge = badge.strip()

    # Ещё строкой выше — название секции
    section = g(f"A{header_row - 2}")
    if isinstance(section, str) and section.strip():
        spec.section_title = section.strip()

    items_first = header_row + 1

    # --- Позиции: строка считается позицией, если есть номер и описание
    for r in range(items_first, SCAN_LIMIT):
        if r in hidden_rows:
            continue
        num, desc = g(f"{col_n}{r}"), g(f"{col_desc}{r}")
        if num is None or desc is None or not str(desc).strip():
            continue
        lines = str(desc).strip().splitlines()
        spec.items.append(
            Item(
                n=str(num).strip(),
                brand=str(g(f"{col_brand}{r}") or "").strip(),
                title=lines[0].strip(),
                body="\n".join(lines[1:]).strip(),
                qty=str(g(f"{col_qty}{r}") or "").strip(),
                price=g(f"{col_price}{r}"),
                total=g(f"{col_sum}{r}"),
                photo=photos.get(r),
            )
        )

    last_item_row = items_first
    for r in range(items_first, SCAN_LIMIT):
        if g(f"{col_n}{r}") is not None and g(f"{col_desc}{r}") is not None:
            last_item_row = r

    # --- Итоги: пары "подпись/сумма" ниже позиций. Скрытые строки книги
    #     пропускаем — там внутренние расчёты (комиссии, рабочая скидка).
    for r in range(last_item_row + 1, SCAN_LIMIT):
        if r in hidden_rows:
            continue
        value = g(f"{col_total_value}{r}")
        if not isinstance(value, (int, float)):
            continue
        # Подпись стоит слева от суммы, но в разных шаблонах по-разному:
        # в форме — вплотную, в спецификации HENGE — в самой первой колонке.
        # Поэтому берём ближайшую непустую текстовую ячейку слева.
        label = None
        for c in range(column_index_from_string(col_total_value) - 1, 0, -1):
            candidate = g(f"{get_column_letter(c)}{r}")
            if isinstance(candidate, str) and candidate.strip():
                label = candidate
                break
        if not label:
            continue
        text = label.strip()
        spec.totals.append(
            {
                "label": text,
                "value": value,
                # Главная строка — итог к оплате
                "grand": text.lower().startswith("итого"),
            }
        )

    # --- Условия: берём из книги, приводим регистр; иначе — стандартные
    found: list[str] = []
    note = ""
    for r in range(last_item_row + 1, SCAN_LIMIT):
        if r in hidden_rows:
            continue
        v = g(f"A{r}")
        if not isinstance(v, str) or len(v.strip()) < 15:
            continue
        t = v.strip()
        if t.startswith("*"):
            note = _normalize_term(t.lstrip("* ").strip())
        else:
            found.append(_normalize_term(t))
    spec.terms = found or DEFAULT_TERMS
    spec.terms_note = note or DEFAULT_TERMS_NOTE

    if not spec.items:
        spec.warnings.append(
            "Не найдено ни одной позиции — проверьте, что колонки совпадают "
            "с шаблоном (№ в A, производитель в C, описание в E)."
        )
    if not spec.totals:
        spec.warnings.append(f"Не найден блок итогов (суммы в колонке {col_total_value}).")

    return spec
